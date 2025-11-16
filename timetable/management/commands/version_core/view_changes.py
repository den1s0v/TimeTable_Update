import os
import hashlib
import logging
from datetime import datetime, timedelta
from timetable.models import FileVersion, Resource, Storage
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl.comments import Comment
from xls2xlsx import XLS2XLSX
import math

# Создаем логгер для текущего модуля
logger = logging.getLogger(__name__)


class ViewChanges:
    """Класс для подсветки изменений и комментирования измененных ячеек в расписании."""

    # Добавленная переменная класса для хранения истории всех изменений
    all_changes_history = {}

    @staticmethod
    def calculate_file_hash(filename):
        """Вычисляет hash файла."""
        logger.debug(f"Calculating file hash for: {filename}")
        hasher = hashlib.md5()
        with open(filename, "rb") as f:
            buf = f.read()
            hasher.update(buf)
        file_hash = hasher.hexdigest()
        logger.debug(f"File hash calculated: {file_hash}")
        return file_hash

    @staticmethod
    def compare_files(file1, file2, time1, time2):
        """
        Сравнивает два файла и возвращает список различий.
        Возвращает список кортежей с информацией о различиях:
        (sheet_name, row, col, old_value, new_value, change_time)
        """
        logger.info(
            f"Comparing files: {file1} (from {time1}) and {file2} (from {time2})"
        )
        differences = []

        try:
            wb1 = load_workbook(file1, rich_text=True)
            wb2 = load_workbook(file2, rich_text=True)
            logger.debug(f"Workbooks loaded successfully")

            for sheet_name in wb1.sheetnames:
                if sheet_name in wb2.sheetnames:
                    ws1 = wb1[sheet_name]
                    ws2 = wb2[sheet_name]
                    logger.debug(f"Comparing sheet: {sheet_name}")

                    for row in range(1, ws1.max_row + 1):
                        for col in range(1, ws1.max_column + 1):
                            cell1 = ws1.cell(row=row, column=col)
                            cell2 = ws2.cell(row=row, column=col)

                            if cell1.value != cell2.value:
                                change_time = time2 if time2 else datetime.now()
                                differences.append(
                                    (
                                        sheet_name,
                                        row,
                                        col,
                                        cell1.value,
                                        cell2.value,
                                        change_time,
                                    )
                                )

            logger.info(f"Found {len(differences)} differences between files")

        except Exception as e:
            logger.error(
                f"Error comparing files {file1} and {file2}: {str(e)}", exc_info=True
            )

        return differences

    @staticmethod
    def create_comment_text(change_history):
        """Создает текст комментария с историей изменений."""
        logger.debug(f"Creating comment text for {len(change_history)} changes")
        comments = []
        previous_value = None

        for change in sorted(change_history, key=lambda x: x["time"]):
            if change["value"] == previous_value:
                continue

            time_str = change["time"]
            value = str(change["value"]) if change["value"] is not None else "∅"
            comments.append(f"{time_str}: {value}")
            previous_value = change["value"]

        comment_text = "\n".join(comments) if comments else "Нет изменений"
        logger.debug(f"Comment text created: {len(comments)} entries")
        return comment_text

    @staticmethod
    def get_color_gradient(start_color, end_color, ratio):
        """Возвращает промежуточный цвет между start_color и end_color на основе ratio (0-1)"""
        logger.debug(
            f"Calculating color gradient: {start_color} -> {end_color}, ratio: {ratio}"
        )

        def hex_to_rgb(hex_color):
            hex_color = hex_color.lstrip("#")
            return tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))

        def rgb_to_hex(rgb):
            return "#%02x%02x%02x" % tuple(min(255, max(0, int(x))) for x in rgb)

        start_rgb = hex_to_rgb(start_color)
        end_rgb = hex_to_rgb(end_color)

        r = start_rgb[0] + (end_rgb[0] - start_rgb[0]) * ratio
        g = start_rgb[1] + (end_rgb[1] - start_rgb[1]) * ratio
        b = start_rgb[2] + (end_rgb[2] - start_rgb[2]) * ratio

        color = rgb_to_hex((r, g, b))
        logger.debug(f"Calculated color: {color}")
        return color

    @staticmethod
    def highlight_differences(file_path, all_changes, expiration_days=30):
        """
        Применяет подсветку и комментарии к файлу на основе всех изменений.
        all_changes: словарь, где ключи - это (sheet_name, row, col),
                    а значения - список словарей {'value': ..., 'time': ...}
        """
        logger.info(f"Highlighting differences in file: {file_path}")
        logger.info(
            f"Processing {len(all_changes)} changed cells with expiration days: {expiration_days}"
        )

        try:
            wb = load_workbook(file_path, rich_text=True)
            now = datetime.now()
            expiration_date = now - timedelta(days=expiration_days)

            start_color = "#5BED50"  # Красный - свежие изменения
            end_color = "#FFFF00"  # Желтый - старые изменения

            processed_cells = 0
            for (sheet_name, row, col), changes in all_changes.items():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    cell = ws.cell(row=row, column=col)

                    valid_changes = [ch for ch in changes]
                    if not valid_changes:
                        continue

                    valid_changes.sort(key=lambda x: x["time"])

                    comment_text = ViewChanges.create_comment_text(valid_changes)
                    cell.comment = Comment(comment_text, "Change Tracker")

                    latest_change = valid_changes[0]
                    time_format = "%Y-%m-%d %H:%M:%S"
                    change_time = datetime.strptime(latest_change["time"], time_format)
                    time_diff = (now - change_time).total_seconds()
                    max_diff = (now - expiration_date).total_seconds()
                    ratio = min(time_diff / max_diff, 1.0)

                    color = ViewChanges.get_color_gradient(
                        start_color, end_color, ratio
                    )
                    fill = PatternFill(
                        start_color=color[1:], end_color=color[1:], fill_type="solid"
                    )
                    cell.fill = fill
                    processed_cells += 1

            wb.save(file_path)
            logger.info(
                f"Successfully highlighted {processed_cells} cells in {file_path}"
            )

        except Exception as e:
            logger.error(
                f"Error highlighting differences in {file_path}: {str(e)}",
                exc_info=True,
            )
            raise

    @staticmethod
    def compare_all_versions(file_versions_list):
        """Сравнивает все версии файлов и возвращает объединенные изменения."""
        logger.info("Starting comparison of all file versions")
        all_changes = {}

        # Проверка на пустой список
        if not file_versions_list:
            logger.warning("Empty file versions list provided")
            ViewChanges.all_changes_history = all_changes
            return all_changes

        # Сортируем список версий по дате (от старой к новой)
        sorted_versions = sorted(
            file_versions_list,
            key=lambda x: x[1],  # Сортируем по второму элементу кортежа (дате)
        )

        # Преобразуем даты в читаемый формат (необязательно, только для отображения)
        formatted_versions = [
            (path, dt.strftime("%Y-%m-%d %H:%M:%S")) for path, dt in sorted_versions
        ]
        logger.debug(f"Sorted versions: {formatted_versions}")

        logger.info(f"Comparing {len(sorted_versions)} file versions")

        # Сравниваем последовательные пары файлов
        for i in range(len(sorted_versions) - 1):
            file1, time1 = sorted_versions[i]
            file2, time2 = sorted_versions[i + 1]

            logger.debug(
                f"Comparing version {i+1} ({time1}) with version {i+2} ({time2})"
            )

            try:
                # Сравниваем файлы
                differences = ViewChanges.compare_files(file1, file2, time1, time2)
                logger.debug(
                    f"Found {len(differences)} differences between versions {i+1} and {i+2}"
                )

                # Обрабатываем различия
                for diff in differences:
                    sheet_name, row, col, old_val, new_val, change_time = diff
                    key = (sheet_name, row, col)

                    if key not in all_changes:
                        all_changes[key] = []

                    if not all_changes[key] or all_changes[key][-1]["value"] != old_val:
                        all_changes[key].append(
                            {
                                "value": old_val,
                                "time": time1.strftime(
                                    "%Y-%m-%d %H:%M:%S"
                                ),  # Форматируем дату
                            }
                        )

                    all_changes[key].append(
                        {
                            "value": new_val,
                            "time": time2.strftime(
                                "%Y-%m-%d %H:%M:%S"
                            ),  # Форматируем дату
                        }
                    )

            except Exception as e:
                logger.error(
                    f"Error comparing {file1} and {file2}: {str(e)}", exc_info=True
                )
                continue

        logger.info(f"Comparison completed. Total changes found: {len(all_changes)}")
        ViewChanges.all_changes_history = all_changes
        return all_changes

    @staticmethod
    def view_changes(file_versions, output_file, expiration_days):
        """
        Основная функция для просмотра изменений между всеми версиями.
        file_versions: список кортежей (file_path, change_time)
        output_file: путь к файлу, в который будут записаны изменения
        expiration_days: срок годности изменений в днях
        """
        logger.info(
            f"Starting view_changes process. Output file: {output_file}, expiration days: {expiration_days}"
        )

        all_changes = ViewChanges.compare_all_versions(file_versions)

        # Получаем последний файл (с самой новой датой)
        if file_versions:
            sorted_versions = sorted(
                file_versions,
                key=lambda x: x[1],  # Сортируем по второму элементу кортежа (дате)
            )
            # Сортируем по дате (по убыванию) и берем первый элемент
            last_file = sorted_versions[-1][0]
            logger.debug(f"Using last file for highlighting: {last_file}")
        else:
            last_file = None
            logger.warning("No file versions available for processing")

        if last_file:
            ViewChanges.highlight_differences(last_file, all_changes, expiration_days)
            if last_file != output_file:
                logger.info(f"Saving highlighted file to: {output_file}")
                wb = load_workbook(last_file)
                wb.save(output_file)
                logger.info(f"File successfully saved to: {output_file}")
        else:
            logger.error("No valid file found for highlighting changes")

        logger.info("view_changes process completed")
